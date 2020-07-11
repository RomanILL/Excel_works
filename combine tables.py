import openpyxl
import datetime
import os
from glob import glob
import xlrd

'''
rb = xlrd.open_workbook('d:/final.xls',formatting_info=True)
sheet = rb.sheet_by_index(0)
for rownum in range(sheet.nrows):
row = sheet.row_values(rownum)
for c_el in row:
print c_el
'''

def try_float(value):
    if value == None:
        return False
    try:
        float(value)
        return float(value)
    except ValueError:
        return value

def try_int(value):
    if value == None:
        return False
    try:
        int(value)
        return int(value)
    except ValueError:
        return value
'''
# функция для открытия файлов чтения xlsx
def open_file_x(filename):
    #print("попытка открыть файл", filename)
    try:
        excelfile = openpyxl.load_workbook(filename)
    except:
        print("что-то пошло не так, возможно файла не существует")
        input("Чтобы закрыть нажмите Enter...")
        return False
    
    sheets = excelfile.sheetnames
    print("Все листы файла:", excelfile.sheetnames)
    excelfile.active = 0
    sheet = excelfile.active
    print("Выбран лист:", sheet)
    maxrows = sheet.max_row
    maxcols = sheet.max_column
    print("строк всего:", rows, "столбцов всего:",cols)
    return (excelfile, sheet, maxrows, maxcols)
'''

# функция открытия файла чтения xls
def open_file(filename):
    #print("попытка открыть файл", filename)
    try:
        excelfile = xlrd.open_workbook(filename, formatting_info = True)
        print("файл", filename, "открыт")
    except:
        print("что-то пошло не так, возможно файла уже открыт или не существует")
        input("Чтобы закрыть нажмите Enter...")
        return False
    sheet = excelfile.sheet_by_index(0)
    maxrows = sheet.nrows
    maxcols = sheet.ncols
    return (excelfile, sheet, maxrows, maxcols)
    
# функция открытия файла записи xlsx
def write_file(w_filename):
    print("Открываем файл output_kods.xlsx")
    flag = 'существующий файл output_kods.xlsx'
    try:
        exitfile = openpyxl.load_workbook('output_kods.xlsx')
        sheet = exitfile['коды']
    except:
        print("Файл не существует - создаем новый файл 'output_kods.xlsx'")
        flag = 'новый файл output_kods.xlsx'
        # создаем новый excel-файл
        exitfile = openpyxl.Workbook()
        #print(exitfile)
        # добавляем новые листы
        exitfile.create_sheet(title = 'коды', index = 0)
        exitfile.create_sheet(title = 'обработанные', index = 1)
        # получаем лист, с которым будем работать
        sheet = exitfile['коды']
        sheet.append(['Код Фармнет', 'Код Магнит Фарма', "Наименование", "Производитель", "Примерная цена (медиана)", "Дата добавления строки"])
    return (exitfile, sheet, flag)


def find_median(listk):

    listk.sort()
    dlin = len(listk)
    #print(dlin, dlin//2)
    #print(listk)
    #print(listk[dlin//2])
    try:
        if dlin%2 == 0:
            med = (listk[int(dlin//2)]+listk[int(dlin//2)-1])/2
        else:
            med = listk[int(dlin//2)]
        return med
    except:
        return "error"

def sravni(keyS):
    for _ in range(len(mainList)):
        ku = 0
        #print(keyS)
        for ops in range(4):
            if keyS[ops] == mainList[_][ops]:
                ku += 1
        if ku == 4:
            return True
    mainList.append(keyS)
    return False

'''
    Программа соединения файлов excel в один
    все файлы нужно поместить в папку программы
    программа выберет файлы xls и соединит их
    если будут повторы, их не запишут
'''

# создаем дату, которую будем указывать
lst = [str(datetime.date.today().day), str(datetime.date.today().month), str(datetime.date.today().year)]
main_date = ".".join(lst)
'''
a = datetime.date(2012, 7, 21)
print(a.year)
print(a.month)
print(a.day)'''

# счетчик добавленных строк
schet = 0

# открываем/создаем файл, куда будем собирать информацию
outfile, o_sheet, flag = write_file('output_kods.xlsx')
# создадим список кортежей существующих вхождений
out_mrows = o_sheet.max_row
mainList = []
for i in range(2, out_mrows+1):
    key = (try_int(o_sheet.cell(row = i, column = 1).value),
                        str(o_sheet.cell(row = i, column = 2).value),
                        str(o_sheet.cell(row = i, column = 3).value),
                        str(o_sheet.cell(row = i, column = 4).value))
    sravni(key)
        
print("Кол-во существующих в файле связок ключей:", len(mainList))
'''
for i in mainList:
    print(i)
'''
print("_"*20)

# открываем папку назначения, откуда будем брать файлы
# вывести текущую директорию
if not os.path.isdir("input_xls"):
    os.mkdir("input_xls")
    input("Папка 'input_xls' не существовала. Пожалуйста, разместите"+
          "там свои фалы для обработки и нажмите Enter...")

os.chdir("input_xls")

# создаем список файлов для чтения - нет - будем сразу перебирать
# находим файлы.xls

for file_name in glob('*.xls'):
    if file_name == "output_kods.xlsx":
        continue
    # открываем текущий файл
    #return (excelfile, sheet, maxrows, maxcols)
    (inp_file, inp_sheet, inp_mrow, inp_mcols) = open_file(file_name)
    for i_row in range(1, inp_mrow):
        #sheet.cell_value(rowx=0, colx=0)
        # мсобираем кортеж для сравнения
        temp_tuple = (try_int(inp_sheet.cell_value(rowx = i_row, colx = 0)),
                      str(inp_sheet.cell_value(rowx = i_row, colx = 1)),
                      str(inp_sheet.cell_value(rowx = i_row, colx = 2)),
                      str(inp_sheet.cell_value(rowx = i_row, colx = 3)))
        #print(sravni)

        # ищем медианную цену
        medlist = []
        for cols_med in range(4, inp_mcols):
            if inp_sheet.cell_value(rowx = i_row, colx = cols_med) != "" :
                if (try_float(inp_sheet.cell_value(rowx = i_row, colx = cols_med)) != 0):
                    medlist.append(try_float(inp_sheet.cell_value(rowx = i_row, colx = cols_med)))

        median = find_median(medlist)
        #print(medlist)
        #print(median)

            # проверяем есть ли такой в output_kods файле через список
        o_sheet = outfile["коды"]
        if not sravni(temp_tuple):
            # если такой записи еще нет, то добавляем запись в файл и добавляем её в список
            # 'Код Фармнет', 'Код Магнит Фарма', "Наименование", "Производитель", "Цена по рынку (медиана)", "Дата добавления строки"
            # записываем полученные значения с текущей датой в output_kods
            o_sheet.append([try_int(temp_tuple[0]),
                            str(temp_tuple[1]),
                            temp_tuple[2],
                            temp_tuple[3],
                            try_float(median),
                            main_date])
            schet += 1
            # добавляем запись в список - добавляется автоматически в функции
            # mainList.append(str((temp_tuple[0], temp_tuple[1], temp_tuple[2], temp_tuple[3], median, main_date)))
            # проверка закончена
            # недостающие данные уже записаны после проверки в output_kods
            # читаем следующую строку
    # пишем из каких файлов собрано
    o_sheet = outfile["обработанные"]
    o_sheet.append([file_name])
    # не забываем закрыть файл перед открытием следующего
    inp_file.release_resources()
    del inp_file
    # и потом обрабатываем следующий файл (см. for выше)
# сохраняем рабочую книгу
os.chdir("..")
try:
    outfile.save('output_kods.xlsx')
    print('Значения записаны в', flag)
except:
    print('Не удалось сохранить файл, возможно он открыт в другой программе.')
    print('Закройте файл "exitdata.xlsx" и повторите попытку.')
# закрываем собранный новый файл
outfile.close()

print("Программа успешно завершена, добавлено", schet, "уникальных связок")
input('Для заверения нажмите Enter...')
