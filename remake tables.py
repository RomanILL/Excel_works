import openpyxl

def yes_no(keystring):
    yesanswer = ('y', 'yes', "д", "да", "yeap")
    noanswer = ('n', 'no', "н", "нет", "noup")
    keystring = keystring.lower()
    while not (keystring in yesanswer or keystring in noanswer):
        keystring = input("повторите: д/н")
        keystring = keystring.lower()
    if keystring in yesanswer:
        return True
    else:
        return False

def is_float(value):
    if value == None:
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False

def superdata(data_str):
    months = {'января' : "01", 'февраля': "02", 'марта': "03", 'апреля': "04",
              'мая':"05",'июня':"06", 'июля': "07", 'августа': "08", 'сентября': "09",
              'октября': "10",'ноября': "11",'декабря': "12"}
    date_n, month, year = (data_str[:-2].split())
    month = months[month]
    if len(date_n) == 1:
        date_n = "0" + date_n
    return date_n + "." + month + "." + year

def doli(kag, summ, itog):
    if kag == "ИТОГО:":
        return None
    return round(summ/itog, 4)

print("Открываем файл today.xlsx")
try:
    wb = openpyxl.load_workbook('today.xlsx')
except:
    print("что-то пошло не так, возможно файла не существует")
    input("Чтобы закрыть нажмите Enter...")
sheets = wb.sheetnames
print("Все листы файла:", wb.sheetnames)
wb.active = 0
sheet = wb.active
print("Выбран лист:", sheet)
rows = sheet.max_row
cols = sheet.max_column
print("строк всего:", rows, "столбцов всего:",cols)


#присваиваем дату для данного документа
datastring = sheet['B2'].value[8:] #переменная с датой
if not yes_no(input("Период, указанный в файле: \
" + datastring + "Сохранить? д/н >>")):
    datastring = input("Введите дату >>")
datastring = superdata(datastring)
#преобразуем дату


#получаем временный словарь городов
#print("получаем названия городов")
cities = {} #словарь городов {столбец: название}
for i in range(3, cols+1):
    tempvalue = sheet.cell(row = 9, column = i).value
    if tempvalue != "":
        cities[i] = tempvalue

#получаем временный словарь контрагентов
#print("получаем названия контрагентов")
k_agents = {} #словарь контрагентов {столбец: название}
for i in range(10, rows+1):
    tempvalue = sheet.cell(row = i, column = 2).value
    if tempvalue != "":
        k_agents[i] = tempvalue

print("все названия городов")
print(cities)
print("все названия контрагентов")
print(k_agents)

#собираем значения из таблицы
Vsumm = {}
itogo = {}
for i in cities.keys():
    for j in k_agents.keys():
        summ = sheet.cell(row = j, column = i).value
        #print(j,i)
        if is_float(summ) and summ !=0:
            Vsumm[i,j] = float(sheet.cell(row = j, column = i).value)
            if k_agents[j] == 'ИТОГО:':
                itogo[cities[i]] = Vsumm[i,j]
            
#print(Vsumm)

#открываем файл для записи и пишем туда новую табличку
print("Открываем файл exitdata.xlsx")
flag = 'существующий файл exitdata.xlsx'
try:
    wb = openpyxl.load_workbook('exitdata.xlsx')
    sheet = wb['таблица']
except:
    print("что-то пошло не так, возможно файла не существует - создаем новый файл")
    flag = 'новый файл exitdata.xlsx'
    # создаем новый excel-файл
    wb = openpyxl.Workbook()
    # добавляем новый лист
    wb.create_sheet(title = 'таблица', index = 0)
    # получаем лист, с которым будем работать
    sheet = wb['таблица']
    sheet.append([None, 'Дата', "Город", "Контрагент", "Сумма", "Итого по городу", "Доля от итого"])

# Заполняем файл значениями
for kkey in Vsumm:
    '''пусто, Дата, Город, Контрагент,
    Сумма, Итог по городу, Доля'''
    dolya = doli(k_agents[kkey[1]], Vsumm[kkey], itogo[cities[kkey[0]]])
    sheet.append([None, datastring, cities[kkey[0]], k_agents[kkey[1]],
                  Vsumm[kkey], itogo[cities[kkey[0]]], dolya ])
try:
    wb.save('exitdata.xlsx')
    print('Значения добавлены в', flag)
except:
    print('Не удалось сохранить файл, возможно он открыт в другой программе.')
    print('Закройте файл "exitdata.xlsx" и повторите попытку.')
input("Чтобы закрыть программу нажмите Enter...")
























    
